
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_worksheet(filename: str, new_file_name: str = None):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    cell = sheet.cell(2, 3)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3).value
        correct_price = cell * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = correct_price
        print(cell, correct_price)

    values = Reference(sheet,
              min_row=2,
              max_row=4,
              min_col=4,
              max_col=4
              )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "f2")

    wb.save(new_file_name if new_file_name else  filename)


process_worksheet("transactions.xlsx", "transactions2.5.xlsx")

# from pathlib import Path
#
# path = Path()
# global_file = path.glob("*.*py")
# for file in global_file:
#     print(file)
# # path.mkdir()